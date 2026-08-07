import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from pathlib import Path
import logging
import sys
from pathlib import Path
_base_dir = Path(__file__).resolve().parent.parent.parent
if str(_base_dir) not in sys.path:
    sys.path.insert(0, str(_base_dir))
from shared.utils.date_formatting import parse_uk_datetime, format_duration_dhms, format_duration
from shared.utils.excel_styling import style_worksheet

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")



def main():
    base_dir = Path(__file__).resolve().parent.parent
    raw_dir = base_dir / "data" / "raw"
    out_dir = base_dir / "outputs"
    proc_dir = base_dir / "data" / "processed"
    
    out_dir.mkdir(parents=True, exist_ok=True)
    proc_dir.mkdir(parents=True, exist_ok=True)
    
    logging.info("Reading raw Meds cases and SLA CSV files...")
    robot_df = pd.read_csv(raw_dir / "Meds-Robot_Cases_Last_12-Months.csv", encoding="cp1252")
    merge_df = pd.read_csv(raw_dir / "Meds-Merge_Cases_Last_12-Months.csv", encoding="cp1252")
    lock_df = pd.read_csv(raw_dir / "Meds-Lock_Cases_Last_12-Months.csv", encoding="cp1252")
    sla_df = pd.read_csv(raw_dir / "Meds_SLA_Last_12-Months.csv", encoding="utf-8-sig")

    robot_df.insert(0, "Trend", "Robot")
    merge_df.insert(0, "Trend", "Merge")
    lock_df.insert(0, "Trend", "Lock")

    all_cases = pd.concat([robot_df, merge_df, lock_df], ignore_index=True)
    logging.info(f"Loaded {len(all_cases)} total cases across 3 trends (Robot: {len(robot_df)}, Merge: {len(merge_df)}, Lock: {len(lock_df)}).")

    # Clean SLA dataset
    sla_clean = sla_df.drop_duplicates(subset=["task"], keep="first").copy()
    sla_clean.rename(columns={
        "start_time": "sla_start_time",
        "pause_time": "sla_pause_time",
        "end_time": "sla_end_time",
        "business_duration": "business_duration_seconds",
        "duration": "actual_duration_seconds",
        "stage": "sla_stage"
    }, inplace=True)

    # Merge cases with SLA
    merged = pd.merge(all_cases, sla_clean, left_on="number", right_on="task", how="left")

    # Calculate TTR
    created = parse_uk_datetime(merged["sys_created_on"])
    resolved = parse_uk_datetime(merged["resolved_at"])
    merged["ttr_seconds"] = (resolved - created).dt.total_seconds()

    # Convert numeric durations
    merged["business_duration_seconds"] = pd.to_numeric(merged["business_duration_seconds"], errors="coerce")
    merged["actual_duration_seconds"] = pd.to_numeric(merged["actual_duration_seconds"], errors="coerce")

    # Format durations into human-readable text
    merged["ttr_formatted"] = merged["ttr_seconds"].apply(format_duration_dhms)
    merged["business_duration_formatted"] = merged["business_duration_seconds"].apply(format_duration_dhms)
    merged["actual_duration_formatted"] = merged["actual_duration_seconds"].apply(format_duration_dhms)

    # Flag unresolved cases
    merged.loc[resolved.isna(), "ttr_formatted"] = "Unresolved"

    # Column ordering & clear business header names
    col_map = {
        "Trend": "Trend",
        "number": "Case Number",
        "priority": "Priority",
        "state": "State",
        "account": "Account",
        "product": "Product",
        "short_description": "Short Description",
        "sys_created_on": "Sys Created On",
        "resolved_at": "Resolved At",
        "ttr_formatted": "TTR (days, hrs, mins)",
        "business_duration_formatted": "SLA Business Duration (days, hrs, mins)",
        "actual_duration_formatted": "SLA Actual Duration (days, hrs, mins)",
        "ttr_seconds": "TTR (Seconds)",
        "business_duration_seconds": "SLA Business Duration (Seconds)",
        "actual_duration_seconds": "SLA Actual Duration (Seconds)",
        "sla_stage": "SLA Stage",
        "sla_start_time": "SLA Start Time",
        "sla_pause_time": "SLA Pause Time",
        "sla_end_time": "SLA End Time",
        "assignment_group": "Assignment Group",
        "assigned_to": "Assigned To",
        "sys_updated_on": "Sys Updated On",
        "sys_updated_by": "Sys Updated By",
        "sys_mod_count": "Sys Mod Count",
        "u_problem": "Problem",
        "description": "Description",
        "close_notes": "Close Notes"
    }

    final_df = merged[list(col_map.keys())].rename(columns=col_map)

    # 1. Save Combined CSV
    combined_csv_path = out_dir / "Meds_Combined_Cases_With_SLA.csv"
    proc_csv_path = proc_dir / "Meds_Combined_Cases_With_SLA.csv"
    final_df.to_csv(combined_csv_path, index=False, encoding="utf-8-sig")
    final_df.to_csv(proc_csv_path, index=False, encoding="utf-8-sig")
    logging.info(f"Saved combined CSV to {combined_csv_path}")

    # 2. Save Master Excel Workbook with individual sheets per trend + Combined
    master_wb_path = out_dir / "Meds_Cases_Trend_Workbooks_With_SLA.xlsx"
    with pd.ExcelWriter(master_wb_path, engine="openpyxl") as writer:
        # Write sheets
        for sheet_name in ["Combined", "Robot", "Merge", "Lock"]:
            if sheet_name == "Combined":
                sheet_df = final_df
            else:
                sheet_df = final_df[final_df["Trend"] == sheet_name]
            sheet_df.to_excel(writer, sheet_name=sheet_name, index=False)
            
    # Apply openpyxl styling to master workbook
    wb = openpyxl.load_workbook(master_wb_path)
    for sheet_name in ["Combined", "Robot", "Merge", "Lock"]:
        ws = wb[sheet_name]
        sheet_df = final_df if sheet_name == "Combined" else final_df[final_df["Trend"] == sheet_name]
        style_worksheet(ws, sheet_df)
    wb.save(master_wb_path)
    logging.info(f"Saved styled Master Excel Workbook to {master_wb_path}")

    # 3. Save Individual Trend Workbooks & CSVs
    for trend in ["Robot", "Merge", "Lock"]:
        trend_df = final_df[final_df["Trend"] == trend]
        
        # Individual CSV
        trend_csv_path = out_dir / f"Meds_{trend}_Cases_With_SLA.csv"
        trend_df.to_csv(trend_csv_path, index=False, encoding="utf-8-sig")
        logging.info(f"Saved {trend} CSV to {trend_csv_path}")
        
        # Individual XLSX Workbook
        trend_xlsx_path = out_dir / f"Meds_{trend}_Cases_With_SLA.xlsx"
        with pd.ExcelWriter(trend_xlsx_path, engine="openpyxl") as writer:
            trend_df.to_excel(writer, sheet_name=f"{trend} Cases", index=False)
            
        wb_single = openpyxl.load_workbook(trend_xlsx_path)
        ws_single = wb_single[f"{trend} Cases"]
        style_worksheet(ws_single, trend_df)
        wb_single.save(trend_xlsx_path)
        logging.info(f"Saved styled {trend} Excel Workbook to {trend_xlsx_path}")

    print("\n--- Summary of Generated Files ---")
    print(f"1. Combined CSV: {combined_csv_path}")
    print(f"2. Master Excel Workbook (Tabs: Combined, Robot, Merge, Lock): {master_wb_path}")
    print(f"3. Individual Trend Workbooks:")
    print(f"   - Robot: {out_dir / 'Meds_Robot_Cases_With_SLA.xlsx'}")
    print(f"   - Merge: {out_dir / 'Meds_Merge_Cases_With_SLA.xlsx'}")
    print(f"   - Lock: {out_dir / 'Meds_Lock_Cases_With_SLA.xlsx'}")
    print(f"4. Individual Trend CSVs:")
    print(f"   - Robot: {out_dir / 'Meds_Robot_Cases_With_SLA.csv'}")
    print(f"   - Merge: {out_dir / 'Meds_Merge_Cases_With_SLA.csv'}")
    print(f"   - Lock: {out_dir / 'Meds_Lock_Cases_With_SLA.csv'}")

if __name__ == "__main__":
    main()
