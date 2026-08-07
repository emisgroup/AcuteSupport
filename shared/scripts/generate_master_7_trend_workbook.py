import pandas as pd
import numpy as np
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path
import logging
import sys
from pathlib import Path
_base_dir = Path(__file__).resolve().parent.parent.parent
if str(_base_dir) not in sys.path:
    sys.path.insert(0, str(_base_dir))
from shared.utils.date_formatting import parse_uk_datetime, format_duration_dhms, format_duration
from shared.utils.excel_styling import style_worksheet

logging.basicConfig(level=logging.INFO, format="%(asctime)s - %(levelname)s - %(message)s")



def process_suite(base_path, cases_files, sla_file, suite_name):
    raw_dir = base_path / "data" / "raw"
    sla_df = pd.read_csv(raw_dir / sla_file, encoding="utf-8-sig")
    
    cases_dfs = []
    for trend_name, fname in cases_files.items():
        df = pd.read_csv(raw_dir / fname, encoding="cp1252")
        df.insert(0, "Trend", trend_name)
        df.insert(0, "Suite", suite_name)
        cases_dfs.append(df)
        
    all_cases = pd.concat(cases_dfs, ignore_index=True)
    
    # SLA cleaning
    sla_clean = sla_df.drop_duplicates(subset=["task"], keep="first").copy()
    sla_clean.rename(columns={
        "start_time": "sla_start_time",
        "pause_time": "sla_pause_time",
        "end_time": "sla_end_time",
        "business_duration": "business_duration_seconds",
        "duration": "actual_duration_seconds",
        "stage": "sla_stage"
    }, inplace=True)

    merged = pd.merge(all_cases, sla_clean, left_on="number", right_on="task", how="left")

    created = parse_uk_datetime(merged["sys_created_on"])
    resolved = parse_uk_datetime(merged["resolved_at"])
    merged["ttr_seconds"] = (resolved - created).dt.total_seconds()

    merged["business_duration_seconds"] = pd.to_numeric(merged["business_duration_seconds"], errors="coerce")
    merged["actual_duration_seconds"] = pd.to_numeric(merged["actual_duration_seconds"], errors="coerce")

    merged["ttr_formatted"] = merged["ttr_seconds"].apply(format_duration_dhms)
    merged["business_duration_formatted"] = merged["business_duration_seconds"].apply(format_duration_dhms)
    merged["actual_duration_formatted"] = merged["actual_duration_seconds"].apply(format_duration_dhms)
    merged.loc[resolved.isna(), "ttr_formatted"] = "Unresolved"

    col_map = {
        "Suite": "Suite",
        "Trend": "Trend",
        "number": "Case Number",
        "priority": "Priority",
        "state": "State",
        "account": "Account",
        "product": "Product",
        "short_description": "Short Description",
        "sys_created_on": "Sys Created On",
        "resolved_at": "Resolved At",
        "ttr_formatted": "TTR (days, hrs, mins)",
        "business_duration_formatted": "SLA Business Duration (days, hrs, mins)",
        "actual_duration_formatted": "SLA Actual Duration (days, hrs, mins)",
        "ttr_seconds": "TTR (Seconds)",
        "business_duration_seconds": "SLA Business Duration (Seconds)",
        "actual_duration_seconds": "SLA Actual Duration (Seconds)",
        "sla_stage": "SLA Stage",
        "sla_start_time": "SLA Start Time",
        "sla_pause_time": "SLA Pause Time",
        "sla_end_time": "SLA End Time",
        "assignment_group": "Assignment Group",
        "assigned_to": "Assigned To",
        "sys_updated_on": "Sys Updated On",
        "sys_updated_by": "Sys Updated By",
        "sys_mod_count": "Sys Mod Count",
        "u_problem": "Problem",
        "description": "Description",
        "close_notes": "Close Notes"
    }

    final_df = merged[list(col_map.keys())].rename(columns=col_map)
    return final_df

def main():
    root_dir = Path(__file__).resolve().parent.parent.parent
    meds_dir = root_dir / "Meds"
    sym_dir = root_dir / "Symphony"
    
    logging.info("Processing Meds cases and SLA...")
    meds_df = process_suite(
        meds_dir,
        {
            "Robot": "Meds-Robot_Cases_Last_12-Months.csv",
            "Merge": "Meds-Merge_Cases_Last_12-Months.csv",
            "Lock": "Meds-Lock_Cases_Last_12-Months.csv"
        },
        "Meds_SLA_Last_12-Months.csv",
        "Meds"
    )

    logging.info("Processing Symphony cases and SLA...")
    sym_df = process_suite(
        sym_dir,
        {
            "Audit": "Sym-Audit_Cases_Last_12-Months.csv",
            "DAD": "Sym-DAD_Cases_Last_12-Months.csv"
        },
        "Sym_SLA_Last_12-Months.csv",
        "Symphony"
    )

    # 7 Worksheets definition:
    # 1. Meds Combined
    # 2. Robot
    # 3. Merge
    # 4. Lock
    # 5. Symphony Combined
    # 6. Audit
    # 7. DAD
    sheets_dict = {
        "Meds Combined": meds_df,
        "Robot": meds_df[meds_df["Trend"] == "Robot"],
        "Merge": meds_df[meds_df["Trend"] == "Merge"],
        "Lock": meds_df[meds_df["Trend"] == "Lock"],
        "Symphony Combined": sym_df,
        "Audit": sym_df[sym_df["Trend"] == "Audit"],
        "DAD": sym_df[sym_df["Trend"] == "DAD"]
    }

    output_filename = "Meds_and_Symphony_Cases_7_Trend_Workbooks_With_SLA.xlsx"
    target_paths = [
        root_dir / output_filename,
        meds_dir / "outputs" / output_filename,
        sym_dir / "outputs" / output_filename
    ]

    for target_path in target_paths:
        target_path.parent.mkdir(parents=True, exist_ok=True)
        with pd.ExcelWriter(target_path, engine="openpyxl") as writer:
            for sname, sdf in sheets_dict.items():
                sdf.to_excel(writer, sheet_name=sname, index=False)
                
        wb = openpyxl.load_workbook(target_path)
        for sname, sdf in sheets_dict.items():
            ws = wb[sname]
            style_worksheet(ws, sdf)
        wb.save(target_path)
        logging.info(f"Saved master 7-sheet workbook to: {target_path}")

    print("\n=== Master Excel Workbook Successfully Generated ===")
    print(f"File location: {target_paths[0]}")
    print("\nWorksheets Included (Total 7):")
    for idx, (sname, sdf) in enumerate(sheets_dict.items(), 1):
        print(f"  {idx}. [{sname}] - {len(sdf)} cases")

if __name__ == "__main__":
    main()
