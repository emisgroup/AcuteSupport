import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

def style_worksheet(ws, df):
    """Apply executive formatting to openpyxl worksheet."""
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=10)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    # Enable autofilter
    ws.auto_filter.ref = ws.dimensions
    
    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Header styling
    for col_num in range(1, len(df.columns) + 1):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[1].height = 28

    # Data row styling & column width calculation
    for row in ws.iter_rows(min_row=2, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
        for cell in row:
            cell.font = data_font
            cell.border = thin_border
            # Alignment rules based on content
            if isinstance(cell.value, (int, float)):
                cell.alignment = Alignment(horizontal="right", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")

    # Column auto-fit width with safety bounds
    for col in ws.columns:
        col_name = str(col[0].value) if col[0].value is not None else ""
        max_len = len(col_name)
        for cell in col[1:100]: # sample first 100 rows for speed
            if cell.value is not None:
                max_len = max(max_len, len(str(cell.value)))
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max(max_len + 3, 12), 50)
